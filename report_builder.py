from __future__ import annotations

import gc
import io
import os
import re
import tempfile
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


REPORT_SHEET_NAME = "Report"
WEBLOG_SHEET_NAME = "Weblog Data"
OUTPUT_FILE_NAME = "Report.xlsx"

FILL_HIGHEST_ACTION = True
FILL_TYPE_COLUMN = True
FILL_IP_COUNT = False

WEBLOG_REQUIRED_COLUMNS = (
    "PARTY_COMPANY_NAME",
    "EMPLOYEE_SHORT_NAME",
    "description",
    "ipAddress",
)
MASTER_REQUIRED_COLUMNS = ("Company Name", "Color")
CRITICAL_REQUIRED_COLUMNS = ("Grp",)
SALES_REQUIRED_COLUMNS = ("Sold Party", "Type", "AMT")

ACTION_COLUMNS = (
    "Detail",
    "Critical_Search",
    "EXCEL",
    "MEDIA",
    "Layout",
    "NEW ARRIVAL",
    "TWIN STONES",
    "WISHLIST",
)

FINAL_COLUMNS = (
    "PARTY_COMPANY_NAME",
    "Type",
    "Remark",
    "Sales_Person",
    "Critical_Search",
    "Detail",
    "EXCEL",
    "MEDIA",
    "Layout",
    "NEW ARRIVAL",
    "TWIN STONES",
    "WISHLIST",
    "IP Counts",
    "Grand Total",
    "AMT",
    "Color",
)

COLOR_MAP = {
    "GREEN": "00FF00",
    "RED": "FF0000",
    "BLUE": "0000FF",
    "YELLOW": "FFFF00",
    "ORANGE": "FFA500",
}


class InputValidationError(ValueError):
    """Raised when an uploaded workbook does not have the expected structure."""


@dataclass(frozen=True)
class BuildResult:
    file_path: str
    file_size: int
    report_rows: int
    source_sheet: str
    warnings: tuple[str, ...]

    def read_bytes(self) -> bytes:
        return Path(self.file_path).read_bytes()

    def cleanup(self) -> None:
        try:
            Path(self.file_path).unlink(missing_ok=True)
        except OSError:
            pass


def _clean_header(value: object) -> str:
    return str(value).strip()


def _read_excel_columns(
    file_bytes: bytes,
    *,
    label: str,
    required: Sequence[str],
    optional: Sequence[str] = (),
    sheet_name: str | int = 0,
) -> pd.DataFrame:
    wanted = set(required) | set(optional)

    try:
        dataframe = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            usecols=lambda column: _clean_header(column) in wanted,
            engine="openpyxl",
        )
    except ValueError as exc:
        raise InputValidationError(f"Unable to read {label}: {exc}") from exc
    except Exception as exc:
        raise InputValidationError(
            f"Unable to open {label}. Confirm that it is a valid .xlsx file."
        ) from exc

    dataframe.columns = [_clean_header(column) for column in dataframe.columns]
    missing = [column for column in required if column not in dataframe.columns]
    if missing:
        raise InputValidationError(
            f"{label} is missing required column(s): {', '.join(missing)}"
        )

    return dataframe


def _find_weblog_sheet(file_bytes: bytes) -> str:
    try:
        with pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl") as workbook:
            sheet_names = workbook.sheet_names
    except Exception as exc:
        raise InputValidationError(
            "Unable to inspect the Weblog workbook. Confirm that it is a valid .xlsx file."
        ) from exc

    if not sheet_names:
        raise InputValidationError("The Weblog workbook does not contain any worksheets.")

    return WEBLOG_SHEET_NAME if WEBLOG_SHEET_NAME in sheet_names else sheet_names[0]


def _normalize_text(series: pd.Series, *, unknown: str | None = None) -> pd.Series:
    if unknown is not None:
        series = series.fillna(unknown)
    else:
        series = series.fillna("")
    return series.astype(str).str.strip().str.upper()


def _normalise_critical_alias(value: object) -> str:
    """Return one cleaned master alias, or an empty string for a blank cell."""
    if pd.isna(value):
        return ""
    alias = str(value).strip().upper()
    return "" if alias == "NAN" else alias


def _build_critical_shapes(critical_filtered: pd.DataFrame) -> tuple[tuple[str, ...], ...]:
    """Build one alias group per critical shape master row.

    `Name` and `Short Name` are aliases for the same shape, so matching both aliases
    in one log description still contributes only one count for that master shape.
    Identical duplicate master rows are also counted only once.
    """
    shapes: list[tuple[str, ...]] = []
    seen: set[frozenset[str]] = set()

    for row in critical_filtered[["Name", "Short Name"]].itertuples(index=False, name=None):
        aliases = frozenset(
            alias
            for alias in (_normalise_critical_alias(value) for value in row)
            if alias
        )
        if not aliases or aliases in seen:
            continue
        seen.add(aliases)
        shapes.append(tuple(sorted(aliases, key=len, reverse=True)))

    return tuple(shapes)


def _critical_match_counts(
    descriptions: pd.Series,
    critical_shapes: Sequence[Sequence[str]],
    eligible_mask: pd.Series,
) -> pd.Series:
    """Count distinct critical master shapes matched in every eligible log row."""
    counts = pd.Series(0, index=descriptions.index, dtype="int64")
    if not critical_shapes or not eligible_mask.any():
        return counts

    eligible_descriptions = descriptions.loc[eligible_mask]
    eligible_counts = pd.Series(0, index=eligible_descriptions.index, dtype="int64")

    for aliases in critical_shapes:
        # Whole-term matching avoids false positives such as OMC inside SOMC.
        pattern = r"(?<!\w)(?:" + "|".join(re.escape(alias) for alias in aliases) + r")(?!\w)"
        shape_matches = eligible_descriptions.str.contains(pattern, regex=True, na=False)
        eligible_counts = eligible_counts.add(shape_matches.astype("int64"), fill_value=0).astype("int64")

    counts.loc[eligible_counts.index] = eligible_counts
    return counts


def _classify_actions(
    logs_df: pd.DataFrame,
    critical_shapes: Sequence[Sequence[str]],
    *,
    critical_search_only_searchid: bool,
) -> tuple[pd.Series, pd.Series]:
    """Return each row's action and the numeric contribution of that row.

    Most log rows contribute 1. A critical-search row contributes the number of
    distinct critical master shapes found in its description.

    When ``critical_search_only_searchid`` is True, critical shapes are counted
    only in descriptions containing ``SEARCHID``. When False, the entire Weblog
    description column is checked for critical shapes.
    """
    descriptions = _normalize_text(logs_df["description"])
    actions = pd.Series("OTHER", index=logs_df.index, dtype="object")
    action_counts = pd.Series(1, index=logs_df.index, dtype="int64")

    searchid_mask = descriptions.str.contains("SEARCHID", regex=False, na=False)
    search_related_mask = searchid_mask | descriptions.str.contains(
        "SEARCH PERFORMED", regex=False, na=False
    )
    actions.loc[search_related_mask] = "Detail"

    if critical_search_only_searchid:
        critical_eligible_mask = searchid_mask
    else:
        critical_eligible_mask = pd.Series(True, index=descriptions.index, dtype="bool")

    critical_counts = _critical_match_counts(
        descriptions,
        critical_shapes,
        critical_eligible_mask,
    )
    critical_mask = critical_counts.gt(0)
    actions.loc[critical_mask] = "Critical_Search"
    action_counts.loc[critical_mask] = critical_counts.loc[critical_mask]

    remaining = ~(search_related_mask | critical_mask)
    media_mask = remaining & descriptions.str.contains(
        r"VIDEO|PHYGITAL|CERTIFICATE|DETAIL|IMAGE|PLOTING", regex=True, na=False
    )
    actions.loc[media_mask] = "MEDIA"

    remaining &= ~media_mask
    ordered_rules: tuple[tuple[str, str], ...] = (
        ("EXCEL", "EXCEL"),
        ("LAYOUT", "Layout"),
        ("NEW ARRIVAL", "NEW ARRIVAL"),
        ("TWIN STONES", "TWIN STONES"),
        ("WISHLIST", "WISHLIST"),
    )

    for keyword, action_name in ordered_rules:
        rule_mask = remaining & descriptions.str.contains(keyword, regex=False, na=False)
        actions.loc[rule_mask] = action_name
        remaining &= ~rule_mask

    return actions, action_counts


def _build_report_dataframe(
    weblog_bytes: bytes,
    master_bytes: bytes,
    critical_bytes: bytes,
    sales_bytes: bytes | None,
    weblog_sheet_name: str,
    *,
    critical_search_only_searchid: bool,
) -> pd.DataFrame:
    logs_df = _read_excel_columns(
        weblog_bytes,
        label="Weblog Data",
        required=WEBLOG_REQUIRED_COLUMNS,
        sheet_name=weblog_sheet_name,
    )
    master_df = _read_excel_columns(
        master_bytes,
        label="Color Master Data",
        required=MASTER_REQUIRED_COLUMNS,
        optional=("Zone",),
    )
    critical_df = _read_excel_columns(
        critical_bytes,
        label="Critical Search Data",
        required=CRITICAL_REQUIRED_COLUMNS,
        optional=("Name", "Short Name"),
    )

    if "Zone" not in master_df.columns:
        master_df["Zone"] = ""
    if "Name" not in critical_df.columns:
        critical_df["Name"] = ""
    if "Short Name" not in critical_df.columns:
        critical_df["Short Name"] = ""

    has_critical_name = (
        critical_df["Name"].fillna("").astype(str).str.strip().ne("").any()
        or critical_df["Short Name"].fillna("").astype(str).str.strip().ne("").any()
    )
    if not has_critical_name:
        raise InputValidationError(
            "Critical Search Data must contain values in 'Name' or 'Short Name'."
        )

    logs_df["PARTY_COMPANY_NAME"] = _normalize_text(logs_df["PARTY_COMPANY_NAME"])
    logs_df["EMPLOYEE_SHORT_NAME"] = _normalize_text(
        logs_df["EMPLOYEE_SHORT_NAME"], unknown="Unknown"
    )
    logs_df["description"] = _normalize_text(logs_df["description"])

    master_df["Company Name"] = _normalize_text(master_df["Company Name"])
    master_df["Color"] = _normalize_text(master_df["Color"])
    master_df["Zone"] = master_df["Zone"].fillna("").astype(str).str.strip()

    critical_df["Grp"] = _normalize_text(critical_df["Grp"])
    critical_filtered = critical_df[critical_df["Grp"].isin(["F3", "CRITICAL"])]
    critical_shapes = _build_critical_shapes(critical_filtered)
    if not critical_shapes:
        raise InputValidationError(
            "Critical Search Data must contain at least one nonblank 'Name' or "
            "'Short Name' under Grp F3 or CRITICAL."
        )

    logs_df["Action"], logs_df["Action_Count"] = _classify_actions(
        logs_df,
        critical_shapes,
        critical_search_only_searchid=critical_search_only_searchid,
    )

    action_counts = (
        logs_df.groupby(
            ["PARTY_COMPANY_NAME", "EMPLOYEE_SHORT_NAME", "Action"],
            dropna=False,
            observed=False,
        )["Action_Count"]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )
    action_counts.columns.name = None

    for column in ACTION_COLUMNS:
        if column not in action_counts.columns:
            action_counts[column] = 0

    action_counts["Grand Total"] = action_counts[list(ACTION_COLUMNS)].sum(axis=1)

    ip_counts = (
        logs_df.groupby(
            ["PARTY_COMPANY_NAME", "EMPLOYEE_SHORT_NAME"],
            dropna=False,
            observed=False,
        )["ipAddress"]
        .nunique(dropna=True)
        .reset_index(name="IP Counts")
    )

    report_df = action_counts.merge(
        ip_counts,
        on=["PARTY_COMPANY_NAME", "EMPLOYEE_SHORT_NAME"],
        how="left",
        validate="one_to_one",
    )
    report_df.rename(columns={"EMPLOYEE_SHORT_NAME": "Sales_Person"}, inplace=True)

    if sales_bytes is not None:
        sales_df = _read_excel_columns(
            sales_bytes,
            label="Sales and Bid Data",
            required=SALES_REQUIRED_COLUMNS,
        )
        sales_df["Sold Party"] = _normalize_text(sales_df["Sold Party"])
        sales_df["Type"] = _normalize_text(sales_df["Type"])
        sales_df["AMT"] = pd.to_numeric(sales_df["AMT"], errors="coerce").fillna(0)

        valid_sales = sales_df[sales_df["Type"].isin(["SALE", "BID"])]
        amount_grouped = (
            valid_sales.groupby("Sold Party", dropna=False, observed=False)["AMT"]
            .sum()
            .reset_index()
        )
        report_df = report_df.merge(
            amount_grouped,
            left_on="PARTY_COMPANY_NAME",
            right_on="Sold Party",
            how="left",
            validate="many_to_one",
        )
        report_df["AMT"] = report_df["AMT"].fillna(0)
    else:
        report_df["AMT"] = 0

    final_df = report_df.merge(
        master_df[["Company Name", "Color", "Zone"]],
        left_on="PARTY_COMPANY_NAME",
        right_on="Company Name",
        how="inner",
        validate="many_to_many",
    )
    final_df.rename(columns={"Zone": "Type"}, inplace=True)
    final_df["Remark"] = ""

    for column in FINAL_COLUMNS:
        if column not in final_df.columns:
            final_df[column] = 0 if column in ACTION_COLUMNS else ""

    final_df = final_df[list(FINAL_COLUMNS)].copy()

    del logs_df, master_df, critical_df, action_counts, ip_counts, report_df
    gc.collect()
    return final_df


def _excel_safe_value(value: object) -> object:
    if value is pd.NA or (not isinstance(value, (str, bytes)) and pd.isna(value)):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _resolve_fill_hex(type_value: object, color_value: object) -> str:
    type_text = str(type_value or "").upper()
    for color_name, hex_code in (
        ("RED", "FF0000"),
        ("ORANGE", "FFA500"),
        ("GREEN", "00FF00"),
        ("YELLOW", "FFFF00"),
        ("BLUE", "0000FF"),
    ):
        if color_name in type_text:
            return hex_code

    color_name = str(color_value or "").upper().strip()
    return COLOR_MAP.get(color_name, "CCCCCC")


def _add_report_sheet(workbook, final_df: pd.DataFrame) -> None:
    if REPORT_SHEET_NAME in workbook.sheetnames:
        del workbook[REPORT_SHEET_NAME]

    worksheet = workbook.create_sheet(REPORT_SHEET_NAME, 0)
    worksheet.sheet_view.showGridLines = True

    for row in dataframe_to_rows(final_df, index=False, header=True):
        worksheet.append([_excel_safe_value(value) for value in row])

    headers = [cell.value for cell in worksheet[1]]
    action_indices = [headers.index(column) + 1 for column in ACTION_COLUMNS]
    ip_index = headers.index("IP Counts") + 1
    color_col_index = headers.index("Color") + 1
    type_col_index = headers.index("Type") + 1
    party_name_index = headers.index("PARTY_COMPANY_NAME") + 1

    fill_cache: dict[str, PatternFill] = {}

    def fill_for(hex_code: str) -> PatternFill:
        if hex_code not in fill_cache:
            fill_cache[hex_code] = PatternFill(
                start_color=hex_code,
                end_color=hex_code,
                fill_type="solid",
            )
        return fill_cache[hex_code]

    special_fill = fill_for("E6B8B7")

    for row_number in range(2, worksheet.max_row + 1):
        party_name = str(
            worksheet.cell(row=row_number, column=party_name_index).value or ""
        ).strip().upper()

        if party_name == "K GIRDHARLAL & CO.":
            for column_number in range(1, color_col_index):
                worksheet.cell(row=row_number, column=column_number).fill = special_fill
            continue

        hex_code = _resolve_fill_hex(
            worksheet.cell(row=row_number, column=type_col_index).value,
            worksheet.cell(row=row_number, column=color_col_index).value,
        )
        fill_style = fill_for(hex_code)

        if FILL_TYPE_COLUMN:
            worksheet.cell(row=row_number, column=type_col_index).fill = fill_style

        if FILL_HIGHEST_ACTION:
            action_values = [
                worksheet.cell(row=row_number, column=index).value or 0
                for index in action_indices
            ]
            maximum = max(action_values) if action_values else 0
            if maximum > 0:
                for index, value in zip(action_indices, action_values):
                    if value == maximum:
                        worksheet.cell(row=row_number, column=index).fill = fill_style

        if FILL_IP_COUNT:
            ip_value = worksheet.cell(row=row_number, column=ip_index).value or 0
            if ip_value >= 5:
                worksheet.cell(row=row_number, column=ip_index).fill = fill_style

    worksheet.delete_cols(color_col_index)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    header_font = Font(bold=True)
    header_fill = fill_for("D9D9D9")

    visible_headers = [cell.value for cell in worksheet[1]]
    amount_index = visible_headers.index("AMT") + 1

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif cell.column == amount_index:
                cell.number_format = "#,##0.00"

    for column_number in range(1, worksheet.max_column + 1):
        maximum_length = 0
        for row_number in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_number, column=column_number).value
            if value is not None:
                maximum_length = max(maximum_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(
            maximum_length + 2,
            60,
        )

    worksheet.row_dimensions[1].height = 22
    workbook.active = 0


def _deduplicate_warnings(messages: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    output: list[str] = []
    for message in messages:
        cleaned = " ".join(str(message).split())
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            output.append(cleaned)
    return tuple(output)


def build_report(
    *,
    weblog_bytes: bytes,
    master_bytes: bytes,
    critical_bytes: bytes,
    sales_bytes: bytes | None = None,
    critical_search_only_searchid: bool = True,
    progress_callback: Callable[[str], None] | None = None,
) -> BuildResult:
    """Generate a report workbook while preserving the original Weblog worksheet."""

    def update(message: str) -> None:
        if progress_callback is not None:
            progress_callback(message)

    update("Checking workbook structure...")
    weblog_sheet_name = _find_weblog_sheet(weblog_bytes)

    update("Reading and aggregating data...")
    final_df = _build_report_dataframe(
        weblog_bytes,
        master_bytes,
        critical_bytes,
        sales_bytes,
        weblog_sheet_name,
        critical_search_only_searchid=critical_search_only_searchid,
    )

    update("Adding the report to the Weblog workbook...")
    warning_messages: list[str] = []
    temporary_path: str | None = None

    try:
        with warnings.catch_warnings(record=True) as captured_warnings:
            warnings.simplefilter("always")
            workbook = load_workbook(
                io.BytesIO(weblog_bytes),
                data_only=False,
                keep_links=False,
            )

            if weblog_sheet_name not in workbook.sheetnames:
                raise InputValidationError(
                    f"The Weblog worksheet '{weblog_sheet_name}' could not be found."
                )

            source_worksheet = workbook[weblog_sheet_name]
            if WEBLOG_SHEET_NAME not in workbook.sheetnames:
                source_worksheet.title = WEBLOG_SHEET_NAME
                source_sheet_output_name = WEBLOG_SHEET_NAME
            else:
                source_sheet_output_name = WEBLOG_SHEET_NAME
            source_worksheet.sheet_state = "visible"

            _add_report_sheet(workbook, final_df)

            update("Saving the completed workbook...")
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary_file:
                temporary_path = temporary_file.name

            workbook.save(temporary_path)
            workbook.close()
            warning_messages.extend(str(item.message) for item in captured_warnings)

        report_rows = len(final_df)
        del final_df, workbook
        gc.collect()

        file_size = Path(temporary_path).stat().st_size
        completed_path = temporary_path
        temporary_path = None
        return BuildResult(
            file_path=completed_path,
            file_size=file_size,
            report_rows=report_rows,
            source_sheet=source_sheet_output_name,
            warnings=_deduplicate_warnings(warning_messages),
        )
    finally:
        if temporary_path and os.path.exists(temporary_path):
            os.remove(temporary_path)
