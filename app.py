import pandas as pd
import numpy as np
import re
import io
import streamlit as st
import threading
import time
import requests
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==========================================
# 0. KEEP-ALIVE BACKGROUND THREAD
# ==========================================
# Replace this with your actual Render URL once you deploy it
APP_URL = "https://your-render-app-name.onrender.com" 

def ping_server():
    """Pings the app every 10 minutes to prevent the server from sleeping."""
    while True:
        try:
            requests.get(APP_URL)
        except Exception:
            pass
        time.sleep(600) 

if 'keep_awake_thread' not in st.session_state:
    thread = threading.Thread(target=ping_server, daemon=True)
    thread.start()
    st.session_state['keep_awake_thread'] = True


# ==========================================
# WORKSHEET COPY HELPER
# ==========================================
def copy_worksheet_as_is(source_ws, target_wb, target_title="Weblog Data"):
    """Copy a worksheet into another workbook while retaining its data and layout."""
    if target_title in target_wb.sheetnames:
        del target_wb[target_title]

    target_ws = target_wb.create_sheet(title=target_title)

    # Copy cell values/formulas and cell-level formatting.
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value

            if source_cell.has_style:
                # Copy style components separately so openpyxl registers them
                # correctly in the destination workbook's style table.
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    # Copy merged cells.
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copy column widths and column visibility/settings.
    for column_letter, source_dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[column_letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed

    # Copy row heights and row visibility/settings.
    for row_number, source_dimension in source_ws.row_dimensions.items():
        target_dimension = target_ws.row_dimensions[row_number]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed

    # Copy common worksheet-level display and print settings.
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal

    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area
    if source_ws.print_title_rows:
        target_ws.print_title_rows = source_ws.print_title_rows
    if source_ws.print_title_cols:
        target_ws.print_title_cols = source_ws.print_title_cols

    # Keep the added sheet visible in the generated report.
    target_ws.sheet_state = 'visible'
    return target_ws

# ==========================================
# 1. UI & CONTROL FLAGS
# ==========================================
st.set_page_config(page_title="Diamond Sales Log Processor", layout="centered")
st.title("💎 Sales Action Report Generator")

# --- CONTROL BUTTONS ---
FILL_HIGHEST_ACTION = True  # If True, highlights the highest count action cells
FILL_TYPE_COLUMN = True      # If True, highlights the 'Type' column
FILL_IP_COUNT = False        # If True, highlights the 'IP Counts' column if >= 5
OUTPUT_FILE_NAME = 'Report.xlsx'
# ----------------------------------------------

# ==========================================
# 2. FILE UPLOADERS
# ==========================================
st.write("Upload your source files below to generate the report.")
weblog_file = st.file_uploader("1. Upload Weblog Data (Excel) *Required*", type=['xlsx'])
master_file = st.file_uploader("2. Upload Color Master Data (Excel) *Required*", type=['xlsx'])
sales_file = st.file_uploader("3. Upload Sales and Bid Data (Excel) *Optional*", type=['xlsx'])
critical_file = st.file_uploader("4. Upload Critical Search Data (Excel) *Required*", type=['xlsx'])

# ==========================================
# 3. GENERATION LOGIC (Fixing the Button Issue)
# ==========================================
# We create the button EXACTLY ONCE here to prevent StreamlitDuplicateElementId errors
if st.button("Generate Report"):
    
    # Check if all required files are present AFTER they click the button
    if weblog_file and master_file and critical_file:
        with st.spinner("Processing data..."):
            try:
                # ==========================================
                # 3. DATA CLEANING & LOADING (with RAM Optimizations)
                # ==========================================
                # Optimized loading: Only pull columns we actually use to prevent RAM crashes
                logs_df = pd.read_excel(weblog_file, usecols=['PARTY_COMPANY_NAME', 'EMPLOYEE_SHORT_NAME', 'description', 'ipAddress'])
                master_df = pd.read_excel(master_file)
                critical_df = pd.read_excel(critical_file)

                # Clean headers
                logs_df.columns = logs_df.columns.str.strip()
                master_df.columns = master_df.columns.str.strip()
                critical_df.columns = critical_df.columns.str.strip()

                if 'Zone' not in master_df.columns:
                    master_df['Zone'] = ''

                logs_df['PARTY_COMPANY_NAME'] = logs_df['PARTY_COMPANY_NAME'].astype(str).str.strip().str.upper()
                logs_df['EMPLOYEE_SHORT_NAME'] = logs_df['EMPLOYEE_SHORT_NAME'].fillna('Unknown').astype(str).str.strip().str.upper()
                logs_df['description'] = logs_df['description'].astype(str).str.strip().str.upper()

                master_df['Company Name'] = master_df['Company Name'].astype(str).str.strip().str.upper()
                master_df['Color'] = master_df['Color'].astype(str).str.strip().str.upper()
                master_df['Zone'] = master_df['Zone'].astype(str).str.strip()

                # ==========================================
                # 4. CRITICAL SEARCH PREP
                # ==========================================
                critical_df['Grp'] = critical_df['Grp'].astype(str).str.strip().str.upper()
                critical_filtered = critical_df[critical_df['Grp'].isin(['F3', 'CRITICAL'])]

                critical_keywords = set(
                    critical_filtered['Name'].dropna().astype(str).str.strip().str.upper().tolist() + 
                    critical_filtered['Short Name'].dropna().astype(str).str.strip().str.upper().tolist()
                )

                def extract_action(desc, critical_keys):
                    if pd.isna(desc) or desc == 'NAN':
                        return 'OTHER'
                    
                    if 'SEARCHID' in desc or 'SEARCH PERFORMED' in desc:
                        for kw in critical_keys:
                            if re.search(r'\b' + re.escape(kw) + r'\b', desc):
                                return 'Critical_Search'
                        return 'Detail' 
                        
                    elif any(keyword in desc for keyword in ['VIDEO', 'PHYGITAL', 'CERTIFICATE', 'DETAIL', 'IMAGE', 'PLOTING']):
                        return 'MEDIA'
                    elif 'EXCEL' in desc:
                        return 'EXCEL'
                    elif 'LAYOUT' in desc:
                        return 'Layout'
                    elif 'NEW ARRIVAL' in desc:
                        return 'NEW ARRIVAL'
                    elif 'TWIN STONES' in desc:
                        return 'TWIN STONES'
                    elif 'WISHLIST' in desc:
                        return 'WISHLIST'
                    
                    return 'OTHER'

                logs_df['Action'] = logs_df['description'].apply(lambda x: extract_action(x, critical_keywords))

                # ==========================================
                # 5. DATA AGGREGATION
                # ==========================================
                action_counts = pd.pivot_table(
                    logs_df, 
                    index=['PARTY_COMPANY_NAME', 'EMPLOYEE_SHORT_NAME'], 
                    columns='Action', 
                    aggfunc='size', 
                    fill_value=0
                ).reset_index()

                required_cols = ['Detail', 'Critical_Search', 'EXCEL', 'MEDIA', 'Layout', 'NEW ARRIVAL', 'TWIN STONES', 'WISHLIST']
                for col in required_cols:
                    if col not in action_counts.columns:
                        action_counts[col] = 0

                action_counts['Grand Total'] = action_counts[required_cols].sum(axis=1)

                ip_counts = logs_df.groupby(['PARTY_COMPANY_NAME', 'EMPLOYEE_SHORT_NAME'])['ipAddress'].nunique().reset_index()
                ip_counts.rename(columns={'ipAddress': 'IP Counts'}, inplace=True)

                report_df = pd.merge(action_counts, ip_counts, on=['PARTY_COMPANY_NAME', 'EMPLOYEE_SHORT_NAME'], how='left')
                report_df.rename(columns={'EMPLOYEE_SHORT_NAME': 'Sales_Person'}, inplace=True)

                # --- OPTIONAL SALES FILE LOGIC ---
                if sales_file is not None:
                    sales_df = pd.read_excel(sales_file, usecols=['Sold Party', 'Type', 'AMT'])
                    sales_df.columns = sales_df.columns.str.strip()
                    sales_df['Sold Party'] = sales_df['Sold Party'].astype(str).str.strip().str.upper()
                    sales_df['Type'] = sales_df['Type'].astype(str).str.strip().str.upper()
                    sales_df['AMT'] = pd.to_numeric(sales_df['AMT'], errors='coerce').fillna(0)

                    valid_sales = sales_df[sales_df['Type'].isin(['SALE', 'BID'])]
                    amt_grouped = valid_sales.groupby('Sold Party')['AMT'].sum().reset_index()

                    report_df = pd.merge(report_df, amt_grouped, left_on='PARTY_COMPANY_NAME', right_on='Sold Party', how='left')
                    report_df['AMT'] = report_df['AMT'].fillna(0)
                else:
                    # If no sales file is uploaded, default the AMT column to 0 for everyone
                    report_df['AMT'] = 0

                # ==========================================
                # 6. MASTER LIST FILTERING & NEW COLUMNS
                # ==========================================
                final_df = pd.merge(
                    report_df, 
                    master_df[['Company Name', 'Color', 'Zone']], 
                    left_on='PARTY_COMPANY_NAME', 
                    right_on='Company Name', 
                    how='inner' 
                )

                final_df.rename(columns={'Zone': 'Type'}, inplace=True)
                final_df['Remark'] = ''

                final_columns = ['PARTY_COMPANY_NAME', 'Type', 'Remark', 'Sales_Person','Critical_Search',  'Detail', 'EXCEL', 'MEDIA', 
                                 'Layout', 'NEW ARRIVAL', 'TWIN STONES', 'WISHLIST', 'IP Counts', 'Grand Total', 'AMT', 'Color']
                final_df = final_df[final_columns]

                # Save to an in-memory buffer
                output = io.BytesIO()
                final_df.to_excel(output, index=False)
                output.seek(0)

                # ==========================================
                # 7. CONDITIONAL FORMATTING & STYLING
                # ==========================================
                COLOR_MAP = {
                    'GREEN': '00FF00', 'RED': 'FF0000', 'BLUE': '0000FF', 
                    'YELLOW': 'FFFF00', 'ORANGE': 'FFA500'
                }

                wb = load_workbook(output)
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                action_indices = [headers.index(col) + 1 for col in required_cols] 
                ip_index = headers.index('IP Counts') + 1
                color_col_index = headers.index('Color') + 1
                type_col_index = headers.index('Type') + 1 
                party_name_index = headers.index('PARTY_COMPANY_NAME') + 1

                for row_num in range(2, ws.max_row + 1):
                    party_name = str(ws.cell(row=row_num, column=party_name_index).value).strip().upper()
                    
                    if party_name == "K GIRDHARLAL & CO.":
                        special_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type='solid')
                        for col_idx in range(1, len(headers)): 
                            ws.cell(row=row_num, column=col_idx).fill = special_fill
                        continue 

                    type_val = str(ws.cell(row=row_num, column=type_col_index).value).upper()
                    
                    if 'RED' in type_val: hex_code = 'FF0000'
                    elif 'ORANGE' in type_val: hex_code = 'FFA500'
                    elif 'GREEN' in type_val: hex_code = '00FF00'
                    elif 'YELLOW' in type_val: hex_code = 'FFFF00'
                    elif 'BLUE' in type_val: hex_code = '0000FF'
                    else:
                        color_name = str(ws.cell(row=row_num, column=color_col_index).value).upper().strip()
                        hex_code = COLOR_MAP.get(color_name, 'CCCCCC') 
                    
                    fill_style = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')

                    if FILL_TYPE_COLUMN:
                        ws.cell(row=row_num, column=type_col_index).fill = fill_style

                    if FILL_HIGHEST_ACTION:
                        action_values = [ws.cell(row=row_num, column=idx).value or 0 for idx in action_indices]
                        max_val = max(action_values) if action_values else 0
                        if max_val > 0: 
                            for idx, val in zip(action_indices, action_values):
                                if val == max_val:
                                    ws.cell(row=row_num, column=idx).fill = fill_style
                                
                    if FILL_IP_COUNT:
                        ip_val = ws.cell(row=row_num, column=ip_index).value or 0
                        if ip_val >= 5:
                            ws.cell(row=row_num, column=ip_index).fill = fill_style

                ws.delete_cols(color_col_index)

                ws.freeze_panes = 'A2' 
                center_align = Alignment(horizontal='center', vertical='center')
                thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") 

                headers = [cell.value for cell in ws[1]]

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = center_align
                        cell.border = thin_border
                        if cell.row > 1 and cell.column <= len(headers) and headers[cell.column - 1] == 'AMT':
                             cell.number_format = '#,##0.00'
                        if cell.row == 1:
                            cell.font = header_font
                            cell.fill = header_fill

                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter 
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = (max_length + 2) 

                # ==========================================
                # 8. ADD ORIGINAL WEBLOG DATA SHEET
                # ==========================================
                # Reset the uploaded file pointer because pandas has already read it.
                weblog_file.seek(0)
                weblog_wb = load_workbook(weblog_file, data_only=False)

                # Prefer a source sheet already named "Weblog Data"; otherwise copy
                # the first active sheet and name it "Weblog Data" in the report.
                source_weblog_ws = (
                    weblog_wb['Weblog Data']
                    if 'Weblog Data' in weblog_wb.sheetnames
                    else weblog_wb.active
                )
                copy_worksheet_as_is(source_weblog_ws, wb, target_title='Weblog Data')
                weblog_wb.close()

                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                
                st.success("Report Generated Successfully!")
                st.download_button(
                    label="📥 Download Report.xlsx",
                    data=final_output,
                    file_name=OUTPUT_FILE_NAME,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"An error occurred: {e}")
                
    else:
        # This will safely trigger if the button is clicked, but files are missing
        st.warning("Please upload the required files (Weblog, Color Master, and Critical Search).")
